import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Cashflow Forecast"

# Styles
header_font = Font(bold=True, size=11)
title_font = Font(bold=True, size=13)
section_font = Font(bold=True, size=11, color="FFFFFF")
formula_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
input_fill = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")
section_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
red_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
rub_format = '#,##0" ₽"'

# Column widths
ws.column_dimensions['A'].width = 38
ws.column_dimensions['B'].width = 16
for col_letter in ['C', 'D', 'E', 'F', 'G', 'H']:
    ws.column_dimensions[col_letter].width = 16
ws.column_dimensions['I'].width = 40

# Title
ws['A1'] = "CASHFLOW FORECAST"
ws['A1'].font = title_font
ws['A2'] = "Последнее обновление: 2026-04-06"
ws['A2'].font = Font(italic=True, color="888888")

# Months header
months = ['Апр', 'Май', 'Июн', 'Июл', 'Авг', 'Сен']
month_cols = ['C', 'D', 'E', 'F', 'G', 'H']

# === SECTION: ПРИХОД ===
row = 4
ws[f'A{row}'] = "ПРИХОД (гросс)"
ws[f'B{row}'] = "Остаток долга"
for i, m in enumerate(months):
    ws[f'{month_cols[i]}{row}'] = m
ws[f'I{row}'] = "Примечание"
for col in ['A', 'B'] + month_cols + ['I']:
    ws[f'{col}{row}'].font = section_font
    ws[f'{col}{row}'].fill = section_fill
    ws[f'{col}{row}'].alignment = Alignment(horizontal='center')

# Debtors data
debtors = [
    ("Марк Вершинин", 525000, [0, 87500, 175000, 175000, 87500, 0], "Рассрочка: май 87.5к, потом по 175к"),
    ("Галина Чубукова", 346500, [0, 57000, 115000, 115000, 59500, 0], "Рассрочка: май 57к, потом по 115к"),
    ("Павел Сабиров", 210000, [70000, 140000, 0, 0, 0, 0], "Апр 70к + май 140к"),
    ("Дима Таранов (гросс → 50% Максиму)", 218500, [0, 100000, 100000, 18500, 0, 0], "По 100к/мес с мая"),
    ("Андрей Козлов (гросс → 50% Максиму)", 70000, [0, 70000, 0, 0, 0, 0], "Остаток в мае"),
    ("Аким (нестандартный 35%)", 106750, [50000, 56750, 0, 0, 0, 0], "2 транша: апр 50к + май остаток"),
    ("Антон Соловьев (не подтверждено)", 150000, [25000, 50000, 50000, 25000, 0, 0], "Ждём подтверждения"),
    ("Новые предоплаты", 0, [0, 0, 0, 0, 0, 0], "Добавлять по мере появления"),
]

start_row = 5
taranov_row = None
kozlov_row = None

for idx, (name, debt, amounts, note) in enumerate(debtors):
    r = start_row + idx
    ws[f'A{r}'] = name
    ws[f'B{r}'] = debt
    ws[f'B{r}'].number_format = rub_format

    if "Таранов" in name:
        taranov_row = r
    if "Козлов" in name:
        kozlov_row = r

    for i, amt in enumerate(amounts):
        cell = ws[f'{month_cols[i]}{r}']
        cell.value = amt
        cell.number_format = rub_format
        cell.fill = input_fill
        cell.border = thin_border

    ws[f'I{r}'] = note
    ws[f'I{r}'].font = Font(italic=True, color="666666", size=9)

last_debtor_row = start_row + len(debtors) - 1

# ИТОГО ГРОСС
total_row = last_debtor_row + 1
ws[f'A{total_row}'] = "ИТОГО ГРОСС"
ws[f'A{total_row}'].font = header_font
for i, col in enumerate(month_cols):
    ws[f'{col}{total_row}'] = f"=SUM({col}{start_row}:{col}{last_debtor_row})"
    ws[f'{col}{total_row}'].number_format = rub_format
    ws[f'{col}{total_row}'].font = header_font
    ws[f'{col}{total_row}'].fill = formula_fill
    ws[f'{col}{total_row}'].border = thin_border

# === SECTION: МАКСИМ 50% ===
row = total_row + 2
ws[f'A{row}'] = "РАСХОД: Максим 50%"
for i, m in enumerate(months):
    ws[f'{month_cols[i]}{row}'] = m
for col in ['A'] + month_cols:
    ws[f'{col}{row}'].font = section_font
    ws[f'{col}{row}'].fill = PatternFill(start_color="C0504D", end_color="C0504D", fill_type="solid")
    ws[f'{col}{row}'].alignment = Alignment(horizontal='center')

max_section_start = row + 1
# Taranов 50%
ws[f'A{max_section_start}'] = "Таранов 50%"
for i, col in enumerate(month_cols):
    ws[f'{col}{max_section_start}'] = f"={col}{taranov_row}*0.5"
    ws[f'{col}{max_section_start}'].number_format = rub_format
    ws[f'{col}{max_section_start}'].fill = formula_fill
    ws[f'{col}{max_section_start}'].border = thin_border

# Козлов 50%
ws[f'A{max_section_start+1}'] = "Козлов 50%"
for i, col in enumerate(month_cols):
    ws[f'{col}{max_section_start+1}'] = f"={col}{kozlov_row}*0.5"
    ws[f'{col}{max_section_start+1}'].number_format = rub_format
    ws[f'{col}{max_section_start+1}'].fill = formula_fill
    ws[f'{col}{max_section_start+1}'].border = thin_border

# Итого Максиму
max_total_row = max_section_start + 2
ws[f'A{max_total_row}'] = "Итого Максиму"
ws[f'A{max_total_row}'].font = header_font
for i, col in enumerate(month_cols):
    ws[f'{col}{max_total_row}'] = f"=SUM({col}{max_section_start}:{col}{max_section_start+1})"
    ws[f'{col}{max_total_row}'].number_format = rub_format
    ws[f'{col}{max_total_row}'].font = header_font
    ws[f'{col}{max_total_row}'].fill = red_fill
    ws[f'{col}{max_total_row}'].border = thin_border

# === НЕТТО ===
netto_row = max_total_row + 2
ws[f'A{netto_row}'] = "ПРИХОД НЕТТО (после Максима)"
ws[f'A{netto_row}'].font = Font(bold=True, size=12)
for i, col in enumerate(month_cols):
    ws[f'{col}{netto_row}'] = f"={col}{total_row}-{col}{max_total_row}"
    ws[f'{col}{netto_row}'].number_format = rub_format
    ws[f'{col}{netto_row}'].font = Font(bold=True, size=12)
    ws[f'{col}{netto_row}'].fill = green_fill
    ws[f'{col}{netto_row}'].border = thin_border

# === SECTION: OPEX ===
opex_header_row = netto_row + 2
ws[f'A{opex_header_row}'] = "РАСХОД: OPEX"
for i, m in enumerate(months):
    ws[f'{month_cols[i]}{opex_header_row}'] = m
for col in ['A'] + month_cols:
    ws[f'{col}{opex_header_row}'].font = section_font
    ws[f'{col}{opex_header_row}'].fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
    ws[f'{col}{opex_header_row}'].alignment = Alignment(horizontal='center')

opex_items = [
    ("Подписки", 10000),
    ("Монтаж", 15000),
    ("Помощница", 15000),
    ("Команда (фикс)", 15000),
]

opex_start = opex_header_row + 1
for idx, (name, amount) in enumerate(opex_items):
    r = opex_start + idx
    ws[f'A{r}'] = name
    for col in month_cols:
        ws[f'{col}{r}'] = amount
        ws[f'{col}{r}'].number_format = rub_format
        ws[f'{col}{r}'].fill = input_fill
        ws[f'{col}{r}'].border = thin_border

opex_last = opex_start + len(opex_items) - 1
opex_total_row = opex_last + 1
ws[f'A{opex_total_row}'] = "Итого OPEX"
ws[f'A{opex_total_row}'].font = header_font
for i, col in enumerate(month_cols):
    ws[f'{col}{opex_total_row}'] = f"=SUM({col}{opex_start}:{col}{opex_last})"
    ws[f'{col}{opex_total_row}'].number_format = rub_format
    ws[f'{col}{opex_total_row}'].font = header_font
    ws[f'{col}{opex_total_row}'].fill = red_fill
    ws[f'{col}{opex_total_row}'].border = thin_border

# === НАЛОГ ===
tax_row = opex_total_row + 2
ws[f'A{tax_row}'] = "Налог СЗ 4% (от гросса)"
ws[f'A{tax_row}'].font = header_font
for i, col in enumerate(month_cols):
    ws[f'{col}{tax_row}'] = f"={col}{total_row}*0.04"
    ws[f'{col}{tax_row}'].number_format = rub_format
    ws[f'{col}{tax_row}'].fill = formula_fill
    ws[f'{col}{tax_row}'].border = thin_border

# === ЧИСТЫЙ CASHFLOW ===
cf_row = tax_row + 2
ws[f'A{cf_row}'] = "ЧИСТЫЙ CASHFLOW"
ws[f'A{cf_row}'].font = Font(bold=True, size=13)
for i, col in enumerate(month_cols):
    ws[f'{col}{cf_row}'] = f"={col}{netto_row}-{col}{opex_total_row}-{col}{tax_row}"
    ws[f'{col}{cf_row}'].number_format = rub_format
    ws[f'{col}{cf_row}'].font = Font(bold=True, size=13)
    ws[f'{col}{cf_row}'].fill = green_fill
    ws[f'{col}{cf_row}'].border = thin_border

# === НАКОПИТЕЛЬНО ===
cum_row = cf_row + 1
ws[f'A{cum_row}'] = "Накопительно (старт 734 926)"
ws[f'A{cum_row}'].font = Font(bold=True, italic=True)
ws[f'C{cum_row}'] = f"=734926+C{cf_row}"
ws[f'C{cum_row}'].number_format = rub_format
ws[f'C{cum_row}'].fill = formula_fill
ws[f'C{cum_row}'].border = thin_border
for i in range(1, len(month_cols)):
    col = month_cols[i]
    prev_col = month_cols[i-1]
    ws[f'{col}{cum_row}'] = f"={prev_col}{cum_row}+{col}{cf_row}"
    ws[f'{col}{cum_row}'].number_format = rub_format
    ws[f'{col}{cum_row}'].fill = formula_fill
    ws[f'{col}{cum_row}'].border = thin_border

# === ПРИМЕЧАНИЯ ===
notes_row = cum_row + 3
notes = [
    "ПРИМЕЧАНИЯ:",
    "• Жёлтые ячейки — вводишь руками (суммы платежей, OPEX)",
    "• Серые ячейки — формулы, не трогать",
    "• В начале месяца: сверь факт с Cashflow логом, скорректируй остатки",
    "• Новый оффер → добавь строку в раздел ПРИХОД перед 'Новые предоплаты'",
    "• Если должник Максима → добавь строку в раздел 'Максим 50%'",
    "• Когда месяц прошёл — можно добавить колонку справа для следующего",
]
for idx, note in enumerate(notes):
    r = notes_row + idx
    ws[f'A{r}'] = note
    if idx == 0:
        ws[f'A{r}'].font = header_font
    else:
        ws[f'A{r}'].font = Font(color="666666", size=10)

# Save
output_path = "/Users/aleksejerofeev/mentor_analytics/ops/Cashflow_Forecast.xlsx"
wb.save(output_path)
print(f"Saved to {output_path}")
